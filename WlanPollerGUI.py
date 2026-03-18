# ==========================================================
# WlanPollerGUI.py  (Updated)
# ==========================================================
# Author: generated/fixed by assistant
# Purpose: Complete GUI for CISCO WLAN POLLER/PARSER (PySide6)
# ==========================================================

import os
import sys
import socket
import re
from datetime import datetime
from typing import List, Optional
from pathlib import Path
from time import time
from PySide6.QtWidgets import QHeaderView
from PySide6.QtCore import Qt, QSize, QThread, QTimer, Signal
from PySide6.QtGui import QFont, QPixmap, QIcon
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QPushButton, QTextEdit, QLineEdit,
    QVBoxLayout, QHBoxLayout, QGridLayout, QComboBox, QCheckBox, QListWidget,
    QStackedWidget, QMessageBox, QGroupBox, QTableWidget, QTableWidgetItem,
    QFileDialog, QSizePolicy, QSpacerItem, QFrame, QProgressBar, QFormLayout
)
from PollerEngine import PollerEngine
from PollerEngine import decrypt_value, encrypt_value
APP_NAME = "CISCO WLAN POLLER GUI"
APP_VERSION = "v5.0.4"
try:
    from ApFlashVulnerableChecker import analyze_logs
except ImportError as e:
    print("ApFlashVulnerableChecker import failed:", e)
    analyze_logs = None

try:
    from PollerEngine import PollerEngine
except ImportError as e:
    raise ImportError(f"CRITICAL: Failed to import PollerEngine module: {e}")

def get_app_base_dir() -> Path:
    """
    Returns directory where:
    - WlanPollerGUI.app lives (macOS)
    - WlanPollerGUI.exe lives (Windows)
    - script folder when running source
    """

    if getattr(sys, "frozen", False):

        exe = Path(sys.executable).resolve()

        # macOS bundled app
        if exe.parent.name == "MacOS" and exe.parent.parent.name == "Contents":
            return exe.parents[3]   # <-- outside .app

        return exe.parent

    return Path(__file__).resolve().parent

BASE_DIR = get_app_base_dir()

DATA_DIR = BASE_DIR / "data"
CONFD_DIR = BASE_DIR / "confd"

DATA_DIR.mkdir(parents=True, exist_ok=True)
CONFD_DIR.mkdir(parents=True, exist_ok=True)



CONFIG_FILE = str(CONFD_DIR / "config.ini")

CONFD = str(CONFD_DIR)
# Optional Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font as XLFont, Alignment
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

# ---------------- Visual constants ----------------
SIDEBAR_BG = "#000000"
HERO_START = "#000000"
HERO_END = "#000000"
CARD_BG = "#ffffff"
CARD_BORDER = "#e6e8eb"
ACCENT = "#16a34a"
TEXT_PRIMARY = "#0f1724"
TEXT_MUTED = "#6b7280"

FONT_BODY = QFont("Roboto", 11)
FONT_TITLE = QFont("Roboto", 22, QFont.Weight.Bold)
FONT_CARD_TITLE = QFont("Roboto", 18, QFont.Weight.DemiBold)


def apply_global_style(app: QApplication):
    """
    Apply global QSS. Uses an f-string triple-quoted string so the CSS
    is a proper Python string literal and can reference color constants.
    """
    app.setFont(FONT_BODY)

    qss = f"""
    QWidget {{
        background: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 #f8fafc, stop:1 #eef2f7);
        color: {TEXT_PRIMARY};
        font-family: Roboto, "Open Sans", "Segoe UI", Arial, sans-serif;
    }}

    QLabel#heroTitle {{
        color: white;
    }}

    QListWidget {{
        background: {SIDEBAR_BG};
        color: white;
        border: none;
        padding-top: 10px;
    }}
    QListWidget::item {{
        padding: 12px 18px;
        border-radius: 6px;
    }}
    QListWidget::item:selected {{
        background: #1c1c1c;
    }}

    QGroupBox {{
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 10px;

        margin-top: 20px;
        padding-top: 18px;  

    }}
    QGroupBox::title {{
        subcontrol-origin: margin;
        subcontrol-position: top left;
        left: 16px;
        top: 6px;
        padding: 0px 6px;
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 10px;
    }}

    QPushButton {{
        background: #000000;
        color: white;
        border-radius: 8px;
        padding: 8px 14px;
        min-width: 110px;
    }}
    QPushButton:hover {{
            background: #1c1c1c;
    }}

    QPushButton:pressed {{
        background: #111111;
    }}


    /* ---------------------------
       NAV buttons: black like sidebar
       Mark a button with: btn.setProperty("nav", True)
       --------------------------- */

    QTextEdit, QLineEdit, QComboBox {{
        background: #ffffff;
        border: 1px solid #e6e8eb;
        border-radius: 6px;
        padding: 6px;
    }}

    QTableWidget {{
    background: white;
    border: 1px solid #d1d5db;          /* outer border */
    border-radius: 10px;
    gridline-color: #e5e7eb;
    selection-background-color: #dcfce7;
    selection-color: #000000;
    }}

    QTableWidget::item {{
    border-right: 1px solid #f1f5f9;    /* subtle vertical lines */
    border-bottom: 1px solid #f1f5f9;   /* subtle row lines */
    padding: 10px;
    }}

    QHeaderView::section {{
    background-color: #f3f4f6;
    border-right: 1px solid #e5e7eb;
    border-bottom: 1px solid #d1d5db;
    padding: 8px;
    font-weight: 600;
     }}
    QProgressBar {{
        text-align: center;
        font-weight: 700;
        color: #000000; /* ensure percent text is black */
    }}
    """

    app.setStyleSheet(qss)


def safe_pixmap(path: str, size: Optional[QSize] = None) -> Optional[QPixmap]:
    if not path or not os.path.exists(path):
        return None
    pix = QPixmap(path)
    if size:
        return pix.scaled(size, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
    return pix


def is_ipv4_or_ipv6(addr: str) -> bool:
    addr = addr.strip()
    if not addr:
        return False
    try:
        socket.inet_pton(socket.AF_INET, addr)
        return True
    except Exception:
        pass
    try:
        socket.inet_pton(socket.AF_INET6, addr)
        return True
    except Exception:
        return False


import configparser
from dataclasses import dataclass



# Writable location for macOS bundled app
from pathlib import Path
import sys
import os


class IniStore:
    def __init__(self, path: str):
        self.path = path
        self.cfg = configparser.ConfigParser(interpolation=None)
        if os.path.exists(path):
            self.cfg.read(path)
        if self.cfg.has_option("WLC", "wlcipaddr"):
            val = self.cfg.get("WLC", "wlcipaddr")
            self.cfg.set("WLC", "wlc_ip", val)
            self.cfg.remove_option("WLC", "wlcipaddr")

    def get(self, section: str, key: str, default: str = "") -> str:
        val = self.cfg.get(section, key, fallback=default)

        if "pasw" in key.lower() or "password" in key.lower() or "enable" in key.lower():
            return decrypt_value(val)

        return val

    def bulk_set(self, section: str, data: dict):
        if not self.cfg.has_section(section):
            self.cfg.add_section(section)

        for k, v in data.items():

            if "pasw" in k.lower() or "password" in k.lower() or "enable" in k.lower():

                v = encrypt_value(v)

            self.cfg.set(section, k, v)

    def save(self):
        Path(self.path).parent.mkdir(parents=True, exist_ok=True)
        with open(self.path, "w", encoding="utf-8") as f:
            self.cfg.write(f)


@dataclass
class ApRow:
    ip: str
    model: str
    name: str


# ---------------- Worker ----------------

class PollerWorker(QThread):
    progress = Signal(int)
    log = Signal(str)
    ap_update = Signal(int, str, str, str, str)
    finished_ok = Signal(dict)
    failed = Signal(str)

    def _engine_progress(self, pct):
        self.progress_sig.emit(pct)

    def _engine_log(self, msg):
        self.log_sig.emit(str(msg))

    def _engine_ap_update(self, idx, ip, model, status):
        self.ap_update_sig.emit(idx, ip, model, status)

    def __init__(
            self,
            operation_type: str,
            workflow: str,
            wlc_cmds: List[str],
            ap_cmds: List[str],
            ap_filter_mode: str,
            site_tag: str,
            model_group: str,
            ap_device: str,
            ap_list_file: str = "",
            ap_mode: str = "AP Custom Cmd List",
    ):
        super().__init__()
        self.operation_type = operation_type
        self.workflow = workflow
        self.wlc_cmds = wlc_cmds
        self.ap_cmds = ap_cmds
        self.ap_filter_mode = ap_filter_mode
        self.site_tag = site_tag
        self.model_group = model_group
        self.ap_device = ap_device
        self.ap_list_file = ap_list_file
        self.ap_mode = ap_mode
    def run(self):
        engine = None
        try:
            # create engine inside try/except so creation failures are visible
            try:

                engine = PollerEngine(
                    log_cb=lambda msg: self.log.emit(msg),
                    progress_cb=lambda pct: self.progress.emit(pct),
                    ap_update_cb=lambda i, ip, model, status, name: self.ap_update.emit(i, ip, model, status, name)
                )
                engine.operation = self.operation_type

                # Only pass workflow to engine when WLC is involved
                if self.operation_type == "WLC & AP":
                    engine.workflow = self.workflow
                else:
                    engine.workflow = ""

            except Exception as e:
                try:
                    self.log.emit(f"PollerWorker: engine creation failed: {e}")
                except Exception:
                    pass
                try:
                    self.failed.emit(str(e))
                except Exception:
                    pass
                return

            try:
                self.log.emit("PollerWorker: engine created, starting operation")
            except Exception:
                pass

            start = datetime.now()
            summary = {"start": start, "operation": self.operation_type}

            # --- WLC Only ---
            if self.operation_type == "WLC Only":
                if not self.wlc_cmds:
                    raise ValueError("WLC Cmd List is empty.")

                out = engine.run_wlc_cmds(self.wlc_cmds)
                summary.update({"wlc_output": out})
                summary["end"] = datetime.now()
                self.finished_ok.emit(summary)
                return

            # --- AP Only ---
            if self.operation_type == "AP Only" and self.ap_list_file:

                ap_rows = []

                with open(self.ap_list_file, "r", encoding="utf-8", errors="ignore") as f:
                    for line in f:
                        s = line.strip()
                        if not s:
                            continue

                        parts = [p.strip() for p in (s.split(",") if "," in s else s.split())]

                        ip = parts[0] if len(parts) >= 1 else ""

                        if len(parts) >= 3:
                            model = parts[1]
                            name = " ".join(parts[2:]).strip()
                        elif len(parts) == 2:
                            model = "UNKNOWN"
                            name = parts[1]
                        else:
                            model = "UNKNOWN"
                            name = ""

                        if not name and ip:
                            name = f"AP_{ip.replace('.', '_')}"

                        if ip:
                            ap_rows.append(ApRow(ip=ip, model=model, name=name))

                if not ap_rows:
                    raise ValueError("AP list file is empty.")

                if not self.ap_cmds:
                    raise ValueError("AP Cmd List is empty.")

                self.log.emit(f"[DEBUG] Parsed AP rows: {len(ap_rows)}")
                self.log.emit(f"[DEBUG] AP list file path = {self.ap_list_file}")
                engine.run_ap_poller(ap_rows, self.ap_device, self.ap_cmds, ap_mode=self.ap_mode)

                summary.update({
                    "ap_total": len(ap_rows),
                    "ap_success": getattr(engine, "success", None),
                    "ap_failed": getattr(engine, "failed", None),
                    "data_dir": getattr(engine, "data_dir", ""),
                    "workflow": self.workflow
                })

                # AP COUNT SUMMARY
                self.log.emit("")
                self.log.emit("=" * 56)
                self.log.emit("  AP COUNT SUMMARY")
                self.log.emit("=" * 56)
                self.log.emit(f"  Total APs in file   : {len(ap_rows)}")
                self.log.emit(f"  APs processed       : {len(ap_rows)}")
                self.log.emit(f"  Success             : {engine.success}")
                self.log.emit(f"  Failed              : {engine.failed}")
                self.log.emit("=" * 56)

                # THEN vulnerability analysis
                if self.workflow == "AP Flash Checker" and analyze_logs:
                    self.log.emit("")
                    self.log.emit("=" * 56)
                    self.log.emit("  RUNNING FLASH SUSCEPTIBILITY ANALYSIS...")
                    self.log.emit("  Please wait — scanning AP output logs.")
                    self.log.emit("=" * 56)
                    self.progress.emit(0)
                    vuln_rows, _ = analyze_logs(str(summary["data_dir"]))
                    summary["vulnerable_rows"] = vuln_rows
                    self.log.emit(f"  Susceptibility scan complete. Found: {len(vuln_rows)} Susceptible AP(s)")
                    self.log.emit("=" * 56)
                    self.progress.emit(100)
                summary["end"] = datetime.now()
                self.finished_ok.emit(summary)
                return
            # --- WLC & AP ---
            if self.operation_type == "WLC & AP":
                if self.wlc_cmds:
                    summary["wlc_output"] = engine.run_wlc_cmds(self.wlc_cmds)
                full = engine.fetch_full_ap_list()
                summary["TotalApCnt"] = len(full)
                filtered = full
                if self.ap_filter_mode == "SITE":
                    filtered, total_from_tag = engine.filter_by_site_tag(full, self.site_tag)
                    summary["TotalApCnt"] = total_from_tag
                    summary["SiteTagNameFilter"] = self.site_tag
                elif self.ap_filter_mode == "MODEL":
                    filtered = engine.filter_by_model_group(full, self.model_group)
                    summary["ApFilter"] = self.model_group

                engine.write_filtered_ap_list(filtered)
                if not filtered:
                    raise ValueError("Filtered AP list is empty.")
                if not self.ap_cmds:
                    raise ValueError("AP Cmd List is empty.")
                # Run AP stage ONLY if user actually selected AP operations
                if self.operation_type != "WLC Only":

                    self.log.emit("[WORKER] Starting AP polling stage...")
                    engine.run_ap_poller(filtered, self.ap_device, self.ap_cmds)

                    summary.update({
                        "ap_total": len(filtered),
                        "ap_success": getattr(engine, "success", None),
                        "ap_failed": getattr(engine, "failed", None),
                        "data_dir": getattr(engine, "data_dir", "")
                    })

                else:
                    self.log.emit("[WORKER] WLC Only selected — skipping AP SSH polling")
                    summary.update({
                        "ap_total": 0,
                        "ap_success": 0,
                        "ap_failed": 0,
                        "data_dir": getattr(engine, "data_dir", "")
                    })

                if self.workflow == "AP Flash Checker" and analyze_logs:
                    self.log.emit("")
                    self.log.emit("=" * 56)
                    self.log.emit(" RUNNING FLASH SUSCEPTIBILITY ANALYSIS...")
                    self.log.emit("  Please wait — scanning AP output logs.")
                    self.log.emit("=" * 56)
                    self.progress.emit(0)
                    vuln_rows, _ = analyze_logs(str(summary["data_dir"]))
                    summary["vulnerable_rows"] = vuln_rows
                    self.log.emit(f"  Scan complete. Found: {len(vuln_rows)} susceptible AP(s)")
                    self.log.emit("=" * 56)
                    self.progress.emit(100)

                summary["end"] = datetime.now()
                self.finished_ok.emit(summary)
                return

            raise ValueError("Unknown operation.")


        except Exception as e:

            import traceback

            traceback.print_exc()

            try:

                self.log.emit(f"[ERROR] {str(e)}")

            except Exception:

                pass

            try:

                self.failed.emit(str(e))

            except Exception:

                pass



        finally:
            # cleanup engine if it exposes shutdown/close
            try:
                if engine is not None:
                    if hasattr(engine, "shutdown") and callable(getattr(engine, "shutdown")):
                        try:
                            engine.shutdown()
                        except Exception:
                            pass
                    if hasattr(engine, "close") and callable(getattr(engine, "close")):
                        try:
                            engine.close()
                        except Exception:
                            pass
            except Exception:
                pass


# ---------------- Main Window ----------------
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.run_in_progress = False
        self.setWindowTitle(f"{APP_NAME} {APP_VERSION}")
        self.resize(1200, 820)
        self.setMinimumSize(1200, 820)
        self._init_state()
        self._build_ui()
        # IMPORTANT: fix initial visibility after widgets exist
        QTimer.singleShot(0, self._post_init_layout_fix)

    def _init_state(self):
        self.operation_type = "WLC Only"
        self.workflow = "Custom CLI Commands"
        self.ap_mode = "AP Custom Cmd List"
        self.ap_filter_mode = "NONE"
        self.site_tag = ""
        self.model_group = "All AP Models"
        self.ap_name_map = {}  # NEW: map ip -> ap name for AP Table first column
        self.ap_device = "cos_qca"
        self.operation_type = "WLC & AP"
        self.ap_list_file = ""
        self.ap_list_path = ""
        self.run_count = 0
        self.wlc_cmds: List[str] = []
        self.ap_cmds: List[str] = []
        if IniStore:
            self.ini = IniStore(CONFIG_FILE)
        else:
            self.ini = None

    def _build_ui(self):
        main = QWidget()
        main_layout = QVBoxLayout(main)
        main_layout.setContentsMargins(0, 0, 0, 0)
        self.setCentralWidget(main)

        hero = QWidget()
        hero_layout = QHBoxLayout(hero)
        hero_layout.setContentsMargins(16, 8, 16, 8)
        hero.setStyleSheet("background: #000000;")
        left_logo = QLabel()
        left_pix = safe_pixmap("assets/cisco_logo.png", QSize(100, 36))
        if left_pix:
            left_logo.setPixmap(left_pix)
        hero_layout.addWidget(left_logo, 0, Qt.AlignmentFlag.AlignVCenter)
        title = QLabel("WLAN POLLER GUI")
        title.setObjectName("heroTitle")
        title.setFont(FONT_TITLE)
        title.setStyleSheet("color: white;")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hero_layout.addWidget(title, 1)
        right_logo = QLabel()
        right_pix = safe_pixmap("assets/wlc_9800.png", QSize(120, 36))
        if right_pix:
            right_logo.setPixmap(right_pix)
        hero_layout.addWidget(right_logo, 0, Qt.AlignmentFlag.AlignVCenter)
        main_layout.addWidget(hero)

        content = QWidget()
        content_layout = QHBoxLayout(content)
        content_layout.setContentsMargins(14, 28, 14, 14)
        content_layout.setSpacing(12)
        main_layout.addWidget(content, 1)

        self.sidebar = QListWidget()
        self.sidebar.setFixedWidth(220)
        self.sidebar.setFont(QFont("Roboto", 12))
        steps = [
            "Step1 Operation Type",
            "Step2 Credentials",
            "Step3 Workflow",
            "Step4 CLI Cmd List",
            "Step5 AP Filters",
            "Step6 Preview",
            "Step7 Run/Results",
            "Parser",
        ]
        self.sidebar.addItems(steps)
        # self.sidebar.setEnabled(False)
        self.sidebar.setEnabled(True)
        # Make clicking the side nav change pages: connect the currentRowChanged
        # signal to the existing _goto_step method.
        # _goto_step expects an int index, so connect directly.
        self.sidebar.currentRowChanged.connect(self._goto_step)
        content_layout.addWidget(self.sidebar)

        self.stack = QStackedWidget()
        content_layout.addWidget(self.stack, 1)

        self.stack.addWidget(self._page_step1())
        self.stack.addWidget(self._page_step2())
        self.stack.addWidget(self._page_step3())
        self.stack.addWidget(self._page_step4())
        self.stack.addWidget(self._page_step5())
        self.stack.addWidget(self._page_step6())
        self.stack.addWidget(self._page_step7())
        self.stack.addWidget(self._page_parser())

        self._goto_step(0)
        QTimer.singleShot(0, self._refresh_visibility)

    def _stop_worker(self, timeout_ms: int = 3000):
        """
        Try to stop self.worker cleanly. If it doesn't stop within timeout, terminate it.
        Called during app close to avoid hanging background threads.
        """
        try:
            if not hasattr(self, "worker") or self.worker is None:
                return
            w = self.worker
            # Request interruption if QThread supports it
            try:
                w.requestInterruption()
            except Exception:
                pass
            # Ask QThread to quit (if run() listens for interruption this will help)
            try:
                w.quit()
            except Exception:
                pass
            # Wait for completion briefly
            try:
                w.wait(timeout_ms)
            except Exception:
                pass
            # If still running, force terminate (last resort)
            if getattr(w, "isRunning", lambda: False)():
                try:
                    w.terminate()
                except Exception:
                    pass
        except Exception:
            pass

    def closeEvent(self, event):
        if getattr(self, "run_in_progress", False):
            QMessageBox.warning(
                self,
                "Operation Running",
                "Please wait until the operation completes."
            )
            event.ignore()
            return

        event.accept()

    def _inject_run_preview_into_log(self):
        """
        Inject Step6 preview block at top of Step7 Run Log.
        Production safe: no passwords exposed.
        """

        if not hasattr(self, "run_log"):
            return

        preview_text = ""

        # Get Step6 preview text safely
        if hasattr(self, "preview_text"):
            try:
                preview_text = self.preview_text.toPlainText().strip()
            except Exception:
                preview_text = ""

        if not preview_text:
            preview_text = "Preview not available."

        header_block = []
        header_block.append("=" * 56)
        header_block.append("RUN CONFIGURATION PREVIEW")
        header_block.append("=" * 56)
        header_block.append(preview_text.strip())
        header_block.append("=" * 56)
        header_block.append("STARTING EXECUTION...")
        header_block.append("=" * 56)
        header_block.append("")

        try:
            self.run_log.append("\n".join(header_block))
        except Exception:
            pass

    # ---------------- Pages ----------------
    def _page_step1(self) -> QWidget:
        """
        Step1 page: card with Choose Operation Type + AP upload (card),
        and the Enter Credentials button placed OUTSIDE the card (below it),
        matching Step2 layout.
        """
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)
        lay.setContentsMargins(8, 8, 8, 8)
        # lay.setContentsMargins(12, 12, 12, 12)
        lay.setAlignment(Qt.AlignTop)

        # --- Card ---
        card = QGroupBox("Step1 - Select Operation Type")
        card.setFont(FONT_CARD_TITLE)
        # Make the card only as tall as its contents (prevents it from filling the page)
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

        c_l = QVBoxLayout(card)
        # c_l.setContentsMargins(12, 12, 12, 12)
        c_l.setContentsMargins(25, 25, 25, 25)
        c_l.setSpacing(18)

        # Label + combobox
        lbl = QLabel("Choose Operation Type:")
        lbl.setStyleSheet("padding-top:4px;")
        c_l.addWidget(lbl)

        self.op_dd = QComboBox()
        # keep same 3 choices but default to "WLC & AP"
        self.op_dd.addItems(["WLC Only", "WLC & AP", "AP Only"])
        self.op_dd.currentTextChanged.connect(self._on_operation_change)
        self.op_dd.setCurrentIndex(1)  # index 1 -> "WLC & AP"
        self.op_dd.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.op_dd.setFixedHeight(30)
        c_l.addWidget(self.op_dd)

        # AP upload row (still inside card)
        self.ap_upload_row = QWidget()
        ab = QHBoxLayout(self.ap_upload_row)
        ab.setContentsMargins(0, 0, 0, 0)
        ab.setSpacing(8)
        ab.addWidget(QLabel("Upload AP List File (Format: AP Ip, AP Name)"))

        self.ap_path = QLineEdit()
        self.ap_path.setReadOnly(True)
        ab.addWidget(self.ap_path)
        self.ap_stats = QLabel("")
        self.ap_stats.setStyleSheet("color:#374151; font-weight:600;")


        self.ap_browse = QPushButton("Browse")
        self.ap_browse.setProperty("class", "secondary")
        self.ap_browse.clicked.connect(self._browse_ap_list)
        ab.addWidget(self.ap_browse)

        c_l.addWidget(self.ap_upload_row)
        self.ap_stats = QLabel("")
        self.ap_stats.setAlignment(Qt.AlignLeft)

        self.ap_stats.setStyleSheet("""
        font-weight:600;
        padding-top:4px;
        """)

        c_l.addWidget(self.ap_stats)

        # small breathing room inside the card (no large stretch)
        c_l.addSpacing(6)

        # --- add card to page layout ---
        lay.addWidget(card)

        # --- Controls row (OUTSIDE the card, just like Step2) ---
        controls = QHBoxLayout()
        controls.setContentsMargins(6, 6, 6, 6)
        controls.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))
        c_l.addSpacing(6)

        self.btn_step1_next = QPushButton("Enter Credentials")
        self.btn_step1_next.setProperty("nav", True)  # keep nav styling if desired
        self.btn_step1_next.setFixedHeight(34)
        self.btn_step1_next.clicked.connect(lambda: self._goto_step(1))
        controls.addWidget(self.btn_step1_next)

        # Add controls *below* the card (not inside it)
        lay.addLayout(controls)

        # Refresh state and return widget
        self._refresh_step1()
        return w

    def _page_step2(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setAlignment(Qt.AlignTop)

        card = QGroupBox("Step2 - WLC / AP Details")
        card.setFont(FONT_CARD_TITLE)
        card_v = QVBoxLayout(card)
        card_v.setContentsMargins(12, 12, 12, 12)
        card_v.setSpacing(6)
        self.step2_top_spacer = QSpacerItem(0, 15, QSizePolicy.Minimum, QSizePolicy.Fixed)
        card_v.addItem(self.step2_top_spacer)
        card_v.addSpacing(15)
        # ---------- WLC BLOCK ----------
        self.wlc_block = QWidget()
        wlc_form = QFormLayout(self.wlc_block)

        self.wlc_ip_label = QLabel("WLC IP Address")
        self.wlc_ip = QLineEdit(self.ini.get("WLC", "wlc_ip") if self.ini else "")
        self.wlc_ip.setFixedHeight(30)

        self.wlc_user_label = QLabel("WLC Username")
        self.wlc_user = QLineEdit(self.ini.get("WLC", "wlc_user") if self.ini else "")
        self.wlc_user.setFixedHeight(30)

        self.wlc_pass_label = QLabel("WLC Password")
        self.wlc_pass = QLineEdit(self.ini.get("WLC", "wlc_pasw") if self.ini else "")
        self.wlc_pass.setEchoMode(QLineEdit.Password)
        self.wlc_pass.setFixedHeight(30)

        wlc_form.addRow(self.wlc_ip_label, self.wlc_ip)
        wlc_form.addRow(self.wlc_user_label, self.wlc_user)
        wlc_form.addRow(self.wlc_pass_label, self.wlc_pass)

        card_v.addWidget(self.wlc_block)

        # ---------- AP BLOCK ----------
        self.ap_block = QWidget()
        ap_form = QFormLayout(self.ap_block)

        self.ap_user_label = QLabel("AP Username")
        self.ap_user = QLineEdit(self.ini.get("AP", "ap_user") if self.ini else "")
        self.ap_user.setFixedHeight(30)

        self.ap_pass_label = QLabel("AP Password")
        self.ap_pass = QLineEdit(self.ini.get("AP", "ap_pasw") if self.ini else "")
        self.ap_pass.setEchoMode(QLineEdit.Password)
        self.ap_pass.setFixedHeight(30)

        self.ap_enable_label = QLabel("Enable Password")
        self.ap_enable = QLineEdit(self.ini.get("AP", "ap_enable") if self.ini else "")
        self.ap_enable.setEchoMode(QLineEdit.Password)
        self.ap_enable.setFixedHeight(30)

        ap_form.addRow(self.ap_user_label, self.ap_user)
        ap_form.addRow(self.ap_pass_label, self.ap_pass)
        ap_form.addRow(self.ap_enable_label, self.ap_enable)

        card_v.addWidget(self.ap_block)

        lay.addWidget(card)

        row = QHBoxLayout()
        back_btn = QPushButton("Back")
        back_btn.setProperty("nav", True)
        back_btn.clicked.connect(lambda: self._goto_step(0))

        save_btn = QPushButton("Save")
        save_btn.setProperty("nav", True)
        save_btn.clicked.connect(self._save_creds)

        proceed_btn = QPushButton("Proceed")
        proceed_btn.setProperty("nav", True)
        proceed_btn.clicked.connect(self._step2_proceed)

        row.addWidget(back_btn);
        row.addWidget(save_btn);
        row.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum));
        row.addWidget(proceed_btn)
        lay.addLayout(row)
        return w

    def _page_step3(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(10)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setAlignment(Qt.AlignTop)

        # Card
        card = QGroupBox("Step3 - Choose WorkFlow")
        card.setFont(FONT_CARD_TITLE)
        # Keep card only as tall as its contents
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)

        c_l = QVBoxLayout(card)
        c_l.setContentsMargins(25, 25, 25, 25)
        c_l.setSpacing(25)

        c_l.addWidget(QLabel("Choose a WorkFlow"))

        self.workflow_dd = QComboBox()
        # add items
        self.workflow_dd.addItems(["AP Flash Checker", "Custom CLI Commands"])
        self.workflow_dd.currentTextChanged.connect(self._on_workflow_change)
        self.workflow_dd.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.workflow_dd.setFixedHeight(30)
        c_l.addWidget(self.workflow_dd)

        # small breathing room inside card (no big stretch)
        c_l.addSpacing(6)

        # Add the card to page layout
        lay.addWidget(card)

        # Controls row (OUTSIDE the card, same style as Step1/Step2)
        row = QHBoxLayout()
        row.setContentsMargins(6, 6, 6, 6)
        row.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))

        back = QPushButton("Back")
        back.setProperty("nav", True)
        back.clicked.connect(lambda: self._goto_step(1))  # go back to Step2 Credentials
        row.addWidget(back)

        nextb = QPushButton("Proceed")
        nextb.setProperty("nav", True)
        nextb.clicked.connect(self._step3_proceed)
        row.addWidget(nextb)

        # Align the row the same way as in Step1/Step2: right-aligned
        # (we already added an expanding spacer before the back button)
        lay.addLayout(row)

        return w

    def _page_step4(self) -> QWidget:
        from PySide6.QtWidgets import QScrollArea

        outer = QWidget()
        outer_lay = QVBoxLayout(outer)
        outer_lay.setContentsMargins(0, 0, 0, 0)
        outer_lay.setSpacing(4)

        # ── SCROLL AREA wrapping the card ──
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)

        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(8)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setAlignment(Qt.AlignTop)

        card = QGroupBox("Step4 - CLI Cmd List")
        card.setFont(FONT_CARD_TITLE)
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        c_l = QVBoxLayout(card)
        c_l.setContentsMargins(12, 18, 12, 12)
        c_l.setSpacing(10)

        # ── WLC CMD BOX ──────────────────────────────────────
        self.wlc_cmd_box = QTextEdit()
        self.wlc_cmd_box.setPlaceholderText("Enter WLC commands (one per line)")
        self.wlc_cmd_box.setFixedHeight(160)
        self.wlc_cmd_box.setAutoFillBackground(True)
        self.wlc_cmd_box.setStyleSheet(
            "QTextEdit { background-color: #ffffff !important; border: 1px solid #e6e8eb; border-radius: 6px; padding: 6px; }")

        self.wlc_cmd_section = QWidget()
        self.wlc_cmd_section.setStyleSheet("background: #ffffff;")
        self.wlc_cmd_section.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        wlc_layout = QVBoxLayout(self.wlc_cmd_section)
        wlc_layout.setContentsMargins(0, 0, 0, 0)
        wlc_layout.setSpacing(6)

        self.wlc_cmd_label = QLabel("WLC Cmd List")
        self.wlc_cmd_label.setStyleSheet("font-weight:600;")
        wlc_layout.addWidget(self.wlc_cmd_label)
        wlc_layout.addWidget(self.wlc_cmd_box)
        c_l.addWidget(self.wlc_cmd_section)

        # ── AP SECTION ───────────────────────────────────────
        self.ap_cmd_section = QWidget()
        self.ap_cmd_section.setStyleSheet("background: #ffffff;")
        self.ap_cmd_section.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Maximum)
        ap_section_layout = QVBoxLayout(self.ap_cmd_section)
        ap_section_layout.setContentsMargins(0, 0, 0, 0)
        ap_section_layout.setSpacing(8)

        ap_cmd_label = QLabel("AP Cmd List:")
        ap_cmd_label.setStyleSheet("font-weight: bold; margin-top: 4px;")
        ap_section_layout.addWidget(ap_cmd_label)

        ap_mode_row = QHBoxLayout()
        ap_mode_row.addWidget(QLabel("AP Mode:"))
        self.ap_mode_dd = QComboBox()
        self.ap_mode_dd.addItems(["AP Custom Cmd List", "AP Image Download"])
        self.ap_mode_dd.currentTextChanged.connect(self._on_ap_mode_changed)
        ap_mode_row.addWidget(self.ap_mode_dd)
        ap_mode_row.addStretch()
        ap_section_layout.addLayout(ap_mode_row)

        self.ap_cmd_box = QTextEdit()
        self.ap_cmd_box.setPlaceholderText("Enter AP CLI commands (one per line)")
        self.ap_cmd_box.setFixedHeight(160)
        self.ap_cmd_box.setStyleSheet(
            "QTextEdit { background: #ffffff; border: 1px solid #e6e8eb; border-radius: 6px; }")
        ap_section_layout.addWidget(self.ap_cmd_box)

        # ── IMAGE DOWNLOAD SETTINGS ──────────────────────────
        self.ftp_group = QGroupBox("AP Image Download Settings")
        self.ftp_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        ftp_layout = QFormLayout()
        ftp_layout.setSpacing(12)
        ftp_layout.setContentsMargins(12, 16, 12, 16)
        self.ftp_group.setLayout(ftp_layout)

        self.proto_dd = QComboBox()
        self.proto_dd.addItems(["TFTP", "SFTP"])
        self.proto_dd.setFixedHeight(30)
        self.proto_dd.currentTextChanged.connect(self._on_proto_changed)
        ftp_layout.addRow("Protocol:", self.proto_dd)

        self.ftp_user_label = QLabel("SFTP Username:")
        self.ftp_user = QLineEdit()
        self.ftp_user.setPlaceholderText("SFTP username")
        self.ftp_user.setFixedHeight(30)
        self.ftp_user.setVisible(False)
        self.ftp_user_label.setVisible(False)

        self.ftp_pasw_label = QLabel("SFTP Password:")
        self.ftp_pasw = QLineEdit()
        self.ftp_pasw.setPlaceholderText("SFTP password")
        self.ftp_pasw.setEchoMode(QLineEdit.Password)
        self.ftp_pasw.setFixedHeight(30)
        self.ftp_pasw.setVisible(False)
        self.ftp_pasw_label.setVisible(False)

        ftp_layout.addRow(self.ftp_user_label, self.ftp_user)
        ftp_layout.addRow(self.ftp_pasw_label, self.ftp_pasw)

        self.ftp_addr = QLineEdit()
        self.ftp_path = QLineEdit()

        self.ftp_group.setVisible(False)
        ap_section_layout.addWidget(self.ftp_group)

        c_l.addWidget(self.ap_cmd_section)
        lay.addWidget(card)
        lay.addStretch()

        scroll.setWidget(w)
        outer_lay.addWidget(scroll, 1)

        # ── BUTTONS (outside scroll, always visible) ─────────
        btn_row = QHBoxLayout()
        btn_row.setContentsMargins(8, 6, 8, 12)
        outer_lay.setContentsMargins(0, 0, 0, 0)

        back_btn = QPushButton("Back")
        back_btn.setProperty("nav", True)
        back_btn.clicked.connect(lambda: self._goto_step(2))

        save_btn = QPushButton("Save")
        save_btn.setProperty("nav", True)
        save_btn.clicked.connect(self._step4_save)

        proceed_btn = QPushButton("Proceed")
        proceed_btn.setProperty("nav", True)
        proceed_btn.clicked.connect(self._step4_proceed)

        btn_row.addWidget(back_btn)
        btn_row.addWidget(save_btn)
        btn_row.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))
        btn_row.addWidget(proceed_btn)
        outer_lay.addLayout(btn_row)

        self._on_ap_mode_changed(self.ap_mode_dd.currentText())
        return outer
    def _page_step5(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setSpacing(6)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setAlignment(Qt.AlignTop)

        card = QGroupBox("Step5 - AP Filters")
        card.setFont(FONT_CARD_TITLE)

        c_l = QGridLayout()
        c_l.setContentsMargins(12, 12, 12, 12)
        c_l.setHorizontalSpacing(12)
        c_l.setVerticalSpacing(6)
        c_l.setSpacing(15)

        r = 0
        # note = QLabel("(Only one filter can be active at a time)")
        # note.setStyleSheet(f"color:{TEXT_MUTED};")
        # c_l.addWidget(note, r, 0, 1, 2)
        r += 1

        # --- APs By Model (show first) ---
        self.chk_model = QCheckBox("APs By Model")
        self.model_dd = QComboBox()
        self.model_dd.addItems([
            "All AP Models",
            "AP1852/2802/3802/4802",
            "C9105AX/9115AX/9120AX",
            "C9117AX/9130AX/9136/9124",
            "C9162/9163/9164/9166",
            "C9171/9172/9174/9176/9178/9179"
        ])
        self.model_dd.setFixedHeight(28)
        self.model_dd.setEnabled(False)
        self.chk_model.toggled.connect(lambda on: self.model_dd.setEnabled(on))
        c_l.addWidget(self.chk_model, r, 0)
        c_l.addWidget(self.model_dd, r, 1)
        r += 1

        # separator
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFrameShadow(QFrame.Sunken)
        c_l.addWidget(sep, r, 0, 1, 2)
        r += 1

        # --- APs By SiteTag (show second) ---
        self.chk_site = QCheckBox("APs By SiteTag(optional)")
        self.site_tag_txt = QLineEdit()
        self.site_tag_txt.setPlaceholderText("Enter SiteTag Name")
        self.site_tag_txt.setFixedHeight(28)
        self.site_tag_txt.setEnabled(False)
        self.chk_site.toggled.connect(lambda on: self.site_tag_txt.setEnabled(on))
        c_l.addWidget(self.chk_site, r, 0)
        c_l.addWidget(self.site_tag_txt, r, 1)
        r += 1

        # Ensure only one filter active at a time
        self.chk_site.toggled.connect(self._enforce_one_filter)
        self.chk_model.toggled.connect(self._enforce_one_filter)

        card.setLayout(c_l)
        lay.addWidget(card)

        # Buttons row
        row = QHBoxLayout()
        back = QPushButton("Back")
        back.setProperty("nav", True)
        back.clicked.connect(lambda: self._goto_step(3 if self.workflow == "Custom CLI Commands" else 2))

        prev = QPushButton("Preview")
        prev.setProperty("nav", True)
        prev.clicked.connect(self._step5_preview)

        row.addWidget(back)
        row.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))
        row.addWidget(prev)
        lay.addLayout(row)
        lay.addStretch()

        return w

    def _page_step6(self) -> QWidget:
        w = QWidget()

        lay = QVBoxLayout(w)
        lay.setSpacing(12)
        lay.setContentsMargins(12, 12, 12, 12)

        # -------- CARD --------
        card = QGroupBox("Step6 - Preview")
        card.setFont(FONT_CARD_TITLE)
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        c_l = QVBoxLayout(card)
        c_l.setContentsMargins(18, 18, 18, 18)

        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        self.preview_text.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        c_l.addWidget(self.preview_text)

        lay.addWidget(card, 1)  # <-- IMPORTANT: stretch factor 1

        # -------- BUTTON ROW --------
        row = QHBoxLayout()

        back = QPushButton("Back")
        back.setProperty("nav", True)
        back.clicked.connect(lambda: self._goto_step(4 if self.operation_type != "WLC Only" else 3))

        confirm = QPushButton("Confirm and Start WlanPoller")
        confirm.setProperty("nav", True)
        confirm.clicked.connect(self._start_run)

        row.addWidget(back)
        row.addStretch()
        row.addWidget(confirm)

        lay.addLayout(row)
        lay.addStretch()
        return w

    def _page_step7(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(12, 4, 12, 10)
        lay.setSpacing(4)

        run_header = QLabel("Run Log (CLI Output)")
        run_header.setStyleSheet("font-size:18px; font-weight:700; padding:2px 0;")
        lay.addWidget(run_header)

        self.run_card = QGroupBox()
        rlay = QVBoxLayout(self.run_card)
        rlay.setContentsMargins(8, 2, 8, 8)
        rlay.setSpacing(4)
        self.run_log = QTextEdit()
        self.run_log.setReadOnly(True)
        self.run_log.setMinimumHeight(120)
        self.run_log.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.run_log.setFont(QFont("Courier New", 13))
        rlay.addWidget(self.run_log)
        lay.addWidget(self.run_card, 3)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.progress.setTextVisible(True)
        self.progress.setAlignment(Qt.AlignCenter)
        self.progress.setFixedHeight(30)
        self.progress.setFont(QFont("Roboto", 11, QFont.Weight.Bold))
        self.progress.setFormat("%p%")
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #d1d5db;
                border-radius: 10px;
                background: #f3f6f9;
                color: #000000;
                text-align: center;
                padding: 2px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #34d399, stop:1 #16a34a);
                border-radius: 10px;
            }
        """)
        lay.addWidget(self.progress)

        # ── AP TABLE ──────────────────────────────────────────
        self.ap_section = QWidget()
        ap_layout = QVBoxLayout(self.ap_section)
        ap_layout.setContentsMargins(0, 0, 0, 0)
        ap_layout.setSpacing(4)
        ap_layout.addWidget(QLabel("AP Table"))

        self.ap_table = QTableWidget(0, 4)
        self.ap_table.setHorizontalHeaderLabels(["AP Name", "AP Model", "AP IP", "Status"])
        header = self.ap_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        self.ap_table.setColumnWidth(0, 220)
        self.ap_table.setColumnWidth(1, 160)
        self.ap_table.setColumnWidth(2, 160)
        self.ap_table.setColumnWidth(3, 500)
        self.ap_table.verticalHeader().setVisible(False)
        self.ap_table.setWordWrap(False)
        self.ap_table.setTextElideMode(Qt.ElideRight)
        self.ap_table.verticalHeader().setDefaultSectionSize(34)
        self.ap_table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
        self.ap_table.setAlternatingRowColors(True)
        self.ap_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.ap_table.setShowGrid(False)
        self.ap_table.setMinimumHeight(100)
        ap_layout.addWidget(self.ap_table)
        lay.addWidget(self.ap_section, 5)

        # ── VULNERABLE TABLE ──────────────────────────────────
        self.vuln_section = QWidget()
        vuln_layout = QVBoxLayout(self.vuln_section)
        vuln_layout.setContentsMargins(0, 0, 0, 0)
        vuln_layout.setSpacing(4)
        vuln_layout.addWidget(QLabel("Susceptible APs & Recovery Table"))

        self.vuln_table = QTableWidget(0, 4)
        self.vuln_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.vuln_table.setMinimumHeight(100)
        self.vuln_table.verticalHeader().setDefaultSectionSize(32)
        self.vuln_table.setHorizontalHeaderLabels(["AP Name", "AP Model", "AP IP", "Recovery"])
        header_v = self.vuln_table.horizontalHeader()
        header_v.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header_v.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header_v.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header_v.setSectionResizeMode(3, QHeaderView.Stretch)
        self.vuln_table.setColumnWidth(0, 220)
        self.vuln_table.setColumnWidth(1, 160)
        self.vuln_table.setColumnWidth(2, 160)
        self.vuln_table.setColumnWidth(3, 500)
        self.vuln_table.setShowGrid(False)
        self.vuln_table.setAlternatingRowColors(True)
        # ── explicitly enable vertical scroll ────────────────
        self.vuln_table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.vuln_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        vuln_layout.addWidget(self.vuln_table)
        lay.addWidget(self.vuln_section, 2)

        # ── RESULT SUMMARY ────────────────────────────────────
        self.results_summary = QTextEdit()
        self.results_summary.setReadOnly(True)
        self.results_summary.setMinimumHeight(50)
        lay.addWidget(QLabel("===== RESULT SUMMARY ====="))
        lay.addWidget(self.results_summary, 3)

        # ── ACTION BUTTONS ────────────────────────────────────
        actions = QHBoxLayout()
        self.btn_save_log = QPushButton("Save Run Log")
        self.btn_save_log.setProperty("nav", True)
        self.btn_save_log.clicked.connect(self._save_run_log)

        self.btn_export_vuln = QPushButton("Export Susceptible Table to Excel")
        self.btn_export_vuln.setProperty("nav", True)
        self.btn_export_vuln.clicked.connect(self._export_vuln_table)

        self.btn_view_logs = QPushButton("View Logs (Open Folder)", clicked=self._open_data_folder)
        self.btn_view_logs.setProperty("nav", True)

        self.btn_close = QPushButton("Close")
        self.btn_close.clicked.connect(self.close)
        self.btn_close.setProperty("nav", True)

        actions.addWidget(self.btn_save_log)
        actions.addWidget(self.btn_export_vuln)
        actions.addWidget(self.btn_view_logs)
        actions.addItem(QSpacerItem(10, 10, QSizePolicy.Expanding, QSizePolicy.Minimum))
        actions.addWidget(self.btn_close)
        lay.addLayout(actions)

        return w
    def _open_status_file(self):
        try:
            path = getattr(self, "last_status_file", "")

            if not path:
                QMessageBox.warning(self, "Missing", "Status summary log not available.")
                return

            folder = os.path.dirname(path)

            if not os.path.exists(folder):
                QMessageBox.warning(self, "Missing", "Log folder not found.")
                return

            if sys.platform.startswith("win"):
                os.startfile(folder)
            elif sys.platform == "darwin":
                os.system(f'open "{folder}"')
            else:
                os.system(f'xdg-open "{folder}"')

        except Exception as e:
            QMessageBox.warning(self, "Error", str(e))

    def _update_step7_visibility(self):
        """Show AP related widgets only if AP involved"""
        show_ap_related = self.operation_type in ("AP Only", "WLC & AP")

        if hasattr(self, "ap_cmd_section"):
            self.ap_section.setVisible(show_ap_related)

        if hasattr(self, "vuln_section"):
            self.vuln_section.setVisible(show_ap_related)

        if hasattr(self, "btn_export_vuln"):
            self.btn_export_vuln.setVisible(show_ap_related)

    def _page_parser(self) -> QWidget:
        w = QWidget();
        lay = QVBoxLayout(w)
        lay.addWidget(QLabel("<h2>Parser</h2>"))
        self.parser_mode = QComboBox();
        self.parser_mode.addItems(["WLC files", "AP files"])
        self.parser_pattern = QLineEdit();
        self.parser_pattern.setPlaceholderText("Enter regex or substring")
        btn = QPushButton("Search");
        btn.clicked.connect(self._run_parser);
        self.parser_out = QTextEdit();
        self.parser_out.setReadOnly(True)
        lay.addWidget(self.parser_mode);
        lay.addWidget(self.parser_pattern);
        lay.addWidget(btn);
        lay.addWidget(self.parser_out)
        row = QHBoxLayout()
        row.addStretch()

        self.btn_parser_close = QPushButton("Close")
        self.btn_parser_close.setProperty("nav", True)
        self.btn_parser_close.clicked.connect(self.close)

        row.addWidget(self.btn_parser_close)
        lay.addLayout(row)
        return w

    # ---------------- Actions / Helpers ----------------
    def _goto_step(self, idx: int):
        self.stack.setCurrentIndex(idx)

        # Force sidebar highlight
        self.sidebar.blockSignals(True)
        self.sidebar.setCurrentRow(idx)
        item = self.sidebar.item(idx)
        if item:
            item.setSelected(True)
        self.sidebar.blockSignals(False)

        try:
            self._refresh_visibility()
            self._update_step7_visibility()
        except Exception:
            pass

        if idx == 5:
            try:
                self._fill_preview()
            except Exception as e:
                if hasattr(self, "run_log"):
                    self.run_log.append(f"[DEBUG] preview build failed: {e}")

    def _on_operation_change(self, value: str):
        self.operation_type = value
        if value != "AP Only":
            self.ap_list_file = ""
            self.ap_list_path = ""
            if hasattr(self, "ap_path"):
                self.ap_path.setText("")
        self._refresh_step1()
        self._refresh_visibility()
        self._update_step7_visibility()
    def _refresh_step1(self):

        is_ap_only = (self.operation_type == "AP Only")

        if hasattr(self, "ap_upload_row"):
            self.ap_upload_row.setVisible(is_ap_only)

        # 🔴 CLEAR STATS WHEN NOT AP ONLY
        if hasattr(self, "ap_stats"):
            if not is_ap_only:
                self.ap_stats.clear()

        if hasattr(self, "btn_step1_next"):
            if is_ap_only:
                self.btn_step1_next.setEnabled(bool(self.ap_list_file) or bool(self.ap_list_path))
            else:
                self.btn_step1_next.setEnabled(True)
    def _browse_ap_list(self):

        path, _ = QFileDialog.getOpenFileName(
            self,
            "Upload AP List File (Format: AP Ip, AP Name)",
            "",
            "Text/CSV (*.txt *.csv);;All Files (*.*)"
        )

        if not path:
            return

        total_cnt = 0
        valid_cnt = 0
        invalid_cnt = 0
        duplicate_cnt = 0

        seen_ips = set()

        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            for i, line in enumerate(f, start=1):

                s = line.strip()
                if not s:
                    continue

                total_cnt += 1

                parts = [p.strip() for p in (s.split(",") if "," in s else s.split())]
                ip = parts[0]

                try:
                    socket.inet_pton(socket.AF_INET, ip)

                    if ip in seen_ips:
                        duplicate_cnt += 1
                    else:
                        valid_cnt += 1
                        seen_ips.add(ip)

                except Exception:
                    invalid_cnt += 1

        if valid_cnt == 0:
            QMessageBox.critical(self, "Invalid File", "No valid AP IP addresses found.")
            return

        os.makedirs(CONFD, exist_ok=True)

        self.ap_list_file = path
        self.ap_path.setText(path)
        self.ap_list_path = path

        # -------- COLORED STATUS TEXT --------
        stats_html = f"""
        <span style='color:#2563eb; font-weight:600;'>Total:</span> {total_cnt} |
        <span style='color:#16a34a; font-weight:600;'>Valid:</span> {valid_cnt} |
        <span style='color:#dc2626; font-weight:600;'>Invalid:</span> {invalid_cnt} |
        <span style='color:#f59e0b; font-weight:600;'>Duplicates:</span> {duplicate_cnt}
        """

        if hasattr(self, "ap_stats"):
            self.ap_stats.setText(stats_html)

        QMessageBox.information(
            self,
            "AP List Loaded",
            f"""File loaded successfully.

    Total APs: {total_cnt}
    Valid APs: {valid_cnt}
    Invalid APs: {invalid_cnt}
    Duplicate APs: {duplicate_cnt}
    """
        )

        self._refresh_step1()
    def _save_creds(self):
        if self.operation_type in ("WLC Only", "WLC & AP"):
            wlc_ip = self.wlc_ip.text().strip()
            if not is_ipv4_or_ipv6(wlc_ip):
                QMessageBox.critical(self, "Invalid WLC IP", "Please enter a valid IPv4/IPv6 address for the WLC.");
                return
        if not self.ini:
            QMessageBox.warning(self, "Warning", "INI backend not available.");
            return
        if self.operation_type in ("WLC Only", "WLC & AP"):
            self.ini.bulk_set("WLC", {
                "wlc_ip": self.wlc_ip.text().strip(),
                "wlc_user": self.wlc_user.text().strip(),
                "wlc_pasw": self.wlc_pass.text()
            })
        if self.operation_type in ("WLC & AP", "AP Only"):
            self.ini.bulk_set("AP", {"ap_user": self.ap_user.text().strip(), "ap_pasw": self.ap_pass.text(),
                                     "ap_enable": self.ap_enable.text()})
        self.ini.save();
        print("DbgWpgui: Save Func Written to file : ",CONFD)
        print("DbgWpgui:Executable:", DATA_DIR)
        print("DbgWpgui:Base dir:", BASE_DIR)

        QMessageBox.information(self, "Saved", "Credentials saved to confd/config.ini")

    def _save_creds_silent(self):
        if not self.ini:
            return
        if self.operation_type in ("WLC Only", "WLC & AP"):
            try:
                self.ini.bulk_set("WLC",
                                  {
                                      "wlc_ip": self.wlc_ip.text().strip(),
                                      "wlc_user": self.wlc_user.text().strip(),
                                      "wlc_pasw": self.wlc_pass.text()
                                  })
            except Exception:
                pass
        if self.operation_type in ("WLC & AP", "AP Only"):
            try:
                self.ini.bulk_set("AP", {"ap_user": self.ap_user.text().strip(), "ap_pasw": self.ap_pass.text(),
                                         "ap_enable": self.ap_enable.text()})
            except Exception:
                pass
        try:
            self.ini.save()
        except Exception:
            pass

    def _on_worker_log(self, text):
        """Append worker log and touch last_progress_time so watchdog knows worker is alive."""
        try:
            if hasattr(self, "run_log"):
                self.run_log.append(str(text))
            self.last_progress_time = time()
        except Exception:
            pass

    def _watchdog_check(self):
        """If the worker is running but no progress/log seen in 30 seconds, warn user."""
        try:
            if not hasattr(self, "worker") or self.worker is None:
                if getattr(self, "watchdog_timer", None):
                    self.watchdog_timer.stop()
                return

            # if worker not running anymore, stop watchdog
            try:
                if not getattr(self.worker, "isRunning", lambda: False)():
                    if getattr(self, "watchdog_timer", None):
                        self.watchdog_timer.stop()
                    return
            except Exception:
                pass

            last = getattr(self, "last_progress_time", None)
            timeout_sec = 30

            # Image download can be silent for a long time — give it more room
            ap_cmds = getattr(self, "ap_cmds", [])
            # If user typed commands, always use them regardless of mode dropdown
            if ap_cmds:
                self.ap_mode = "AP Custom Cmd List"
            is_image_run = any(
                "archive download-sw" in c.lower()
                or "sftp://" in c.lower()
                or "scp://" in c.lower()
                for c in ap_cmds
            )
            effective_timeout = 3600 if is_image_run else timeout_sec

            if last is None or (time() - last) > effective_timeout:
                if hasattr(self, "run_log"):
                    self.run_log.append(
                        f"[WATCHDOG] No progress or log for {effective_timeout}s. "
                        f"Worker may be waiting on a long-running device command."
                    )
                # optionally request interruption once
                try:
                    if hasattr(self.worker, "requestInterruption"):
                        self.worker.requestInterruption()
                except Exception:
                    pass
        except Exception:
            pass

    def _start_run(self):
        """
        Start the run without any confirmation dialog.
        Saves credentials silently, resets UI, builds the PollerWorker and starts it.
        """
        print('DbgWpgui: Inside Start_run..')

        # Silent save
        try:
            self._save_creds_silent()
        except Exception:
            pass
        self.ap_name_map = {}
        # Pre-populate name map so AP Only 2-column files (IP Name) display correctly
        if self.operation_type == "AP Only" and getattr(self, "ap_list_file", ""):
            try:
                with open(self.ap_list_file, "r", encoding="utf-8", errors="ignore") as _f:
                    for _line in _f:
                        _s = _line.strip()
                        if not _s:
                            continue
                        _parts = [p.strip() for p in (_s.split(",") if "," in _s else _s.split())]
                        _ip = _parts[0] if len(_parts) >= 1 else ""
                        # 3-col: ip model name → name is parts[2]
                        # 2-col: ip name      → name is parts[1]
                        _name = _parts[2] if len(_parts) >= 3 else (_parts[1] if len(_parts) == 2 else "")
                        if _ip and _name:
                            self.ap_name_map[_ip] = _name
            except Exception:
                pass
        # Reset UI defensively
        for attr, op in (
                ("run_log", lambda w: w.clear()),
                ("ap_table", lambda w: w.setRowCount(0)),
                ("vuln_table", lambda w: w.setRowCount(0)),
                ("progress", lambda w: w.setValue(0)),
                ("results_summary", lambda w: w.clear()),
        ):

            if hasattr(self, attr):
                try:
                    op(getattr(self, attr))
                except Exception:
                    pass
        # Disconnect old worker signals before creating new one
        if hasattr(self, "worker") and self.worker is not None:
            try:
                self.worker.log.disconnect()
                self.worker.progress.disconnect()
                self.worker.ap_update.disconnect()
                self.worker.finished_ok.disconnect()
                self.worker.failed.disconnect()
            except Exception:
                pass
            self.worker = None
        # Decide ap_list_file and ap_cmds
        ap_list_file = getattr(self, "ap_list_file", "")


        if self.operation_type != "AP Only":
            ap_list_file = ""

        if getattr(self, "workflow", "") == "AP Flash Checker":
            ap_cmds = getattr(self, "ap_cmds", []) or [
                "show clock",
                "show version",
                "show flash",
                "show flash | i cnssdaemon.log",
                "show boot",
                "show filesystems",
                "show image integrity",
            ]
        else:
            ap_cmds = getattr(self, "ap_cmds", [])



        # Build the worker
        try:
            self.worker = PollerWorker(
                operation_type=self.operation_type,
                workflow=self.workflow,
                wlc_cmds=getattr(self, "wlc_cmds", []),
                ap_cmds=ap_cmds,
                ap_filter_mode=getattr(self, "ap_filter_mode", "NONE"),
                site_tag=getattr(self, "site_tag", ""),
                model_group=getattr(self, "model_group", "All AP Models"),
                ap_device=getattr(self, "ap_device", "cos_qca"),
                ap_list_file=ap_list_file,
                ap_mode=getattr(self, "ap_mode", "AP Custom Cmd List"),
            )
        except Exception as e:
            QMessageBox.critical(self, "Worker Error", f"Failed to create worker: {e}")
            return

        # Prepare watchdog state
        self.last_progress_time = time()
        if getattr(self, "watchdog_timer", None) is None:
            self.watchdog_timer = QTimer(self)
            self.watchdog_timer.setInterval(10000)  # check every 10 seconds
            self.watchdog_timer.timeout.connect(self._watchdog_check)

        # Hook signals with small wrappers that update last_progress_time
        try:
            # log -> wrapper that updates last_progress_time

            self.worker.log.connect(self._on_worker_log)
        except Exception:
            pass

        try:
            # progress -> update bar and last_progress_time

            def _progress_cb(pct):
                try:
                    if hasattr(self, "progress"):
                        self.progress.setValue(pct)
                    self.last_progress_time = time()
                except Exception:
                    pass

            self.worker.progress.connect(self._ui_progress_update)
        except Exception:
            pass

        try:

            self.worker.ap_update.connect(self._on_ap_update)
        except Exception:
            pass

        # finished: stop watchdog and delegate to your existing _on_finished
        try:
            if hasattr(self.worker, "finished_ok") and callable(getattr(self.worker.finished_ok, "connect", None)):
                def _on_finished_wrapper(summary):
                    try:
                        if getattr(self, "watchdog_timer", None):
                            self.watchdog_timer.stop()
                    except Exception:
                        pass
                    self.run_in_progress = False
                    if hasattr(self, "btn_close"):
                        self.btn_close.setEnabled(True)
                    try:
                        self._on_finished(summary)
                    except Exception:
                        pass

                    # IMPORTANT: re-enable close button (missing piece)
                    try:
                        if hasattr(self, "btn_close"):
                            self.btn_close.setEnabled(True)
                    except Exception:
                        pass

                self.worker.finished_ok.connect(_on_finished_wrapper)
        except Exception:
            pass

        # failed: stop watchdog and show error (also re-enable sidebar)
        try:
            if hasattr(self.worker, "failed") and callable(getattr(self.worker.failed, "connect", None)):
                def _on_fail(e):
                    self.run_in_progress = False

                    try:
                        if getattr(self, "watchdog_timer", None):
                            self.watchdog_timer.stop()
                    except Exception:
                        pass
                    try:
                        if hasattr(self, "run_log"):
                            self.run_log.append("[WORKER FAILED] " + str(e))
                    except Exception:
                        pass
                    try:
                        QMessageBox.critical(self, "Run Failed", e)
                    except Exception:
                        pass
                    try:
                        if hasattr(self, "sidebar"):
                            self.sidebar.setEnabled(True)
                    except Exception:
                        pass

                    if hasattr(self, "btn_close"):
                        self.btn_close.setEnabled(True)
                    self.run_in_progress = False

                self.worker.failed.connect(_on_fail)
        except Exception:
            pass

        # When thread starts -> append a visible message
        try:
            self.worker.started.connect(lambda: self.run_log.append("[WORKER] threads started"))
        except Exception:
            pass

        # Optionally lock navigation / sidebar while run is active
        try:
            if hasattr(self, "sidebar"):
                self.sidebar.setEnabled(False)
        except Exception:
            pass

        # Start watchdog and the worker
        try:
            self.last_progress_time = time()
            self.watchdog_timer.start()
            self._goto_step(6)
            self._inject_run_preview_into_log()
        except Exception:
            try:
                self.stack.setCurrentIndex(6)
            except Exception:
                pass
        # disable close while execution running
        if hasattr(self, "btn_close"):
            self.btn_close.setEnabled(False)
        self.run_in_progress = True
        if self.operation_type == "AP Only":
            # Pre-read file to know row count
            with open(self.ap_list_file, "r", encoding="utf-8", errors="ignore") as f:
                lines = [l for l in f if l.strip()]

        try:
            start_time = datetime.now().strftime("%H:%M:%S")

            msg = (
                "===== STARTING WLAN POLLER =====\n"
                f"Operation: {self.operation_type}\n"
                f"Workflow: {self.workflow}\n"
                f"Start Time: {start_time}\n"
            )
            self.run_log.append(msg)
            self.worker.start()
            if hasattr(self, "run_log"):
                self.run_log.append("[WORKER] start() called")
        except Exception as e:
            QMessageBox.critical(self, "Start Error", f"Failed to start worker: {e}")
            if hasattr(self, "sidebar"):
                self.sidebar.setEnabled(True)

    def _step2_proceed(self):
        if self.operation_type == "WLC Only":
            self._goto_step(3)
        else:
            self._goto_step(2)

    def _step3_proceed(self):
        """
        Proceed from Step3 (Workflow).
        - AP Flash Checker: auto-fill AP commands and SKIP Step4 -> Step5 (Filters).
        - AP Image Download: set ap_mode and SKIP Step4 -> Step5 (user must provide FTP elsewhere).
        - Custom CLI Commands: go to Step4 so user can edit WLC/AP cmd lists.
        """
        # Prefer the widget value if present, else use stored state
        if hasattr(self, "workflow_dd") and callable(getattr(self.workflow_dd, "currentText", None)):
            wf = self.workflow_dd.currentText()
        else:
            wf = getattr(self, "workflow", None)

        # Persist selection in UI state
        self.workflow = wf

        # AP Flash Checker -> set default AP cmds and skip Step4
        if wf == "AP Flash Checker":

            self.ap_cmds = [
                "show clock",
                "show version",
                "show flash",
                "show flash | i cnssdaemon.log",
                "show boot",
                "show filesystems",
                "show image integrity",
            ]

            self.ap_mode = getattr(self, "ap_mode", "AP Custom Cmd List")

            # ---- NEW LOGIC ----
            if self.operation_type == "AP Only":
                # Skip filters → go directly to preview
                self.ap_filter_mode = "NONE"
                self._fill_preview()
                self._goto_step(5)  # Step6 Preview
            else:
                # WLC & AP still uses filters
                self._goto_step(4)

            return

        # AP Image Download -> remember mode and skip Step4
        if wf == "AP Image Download":
            self.ap_mode = "AP Image Download"
            # If you want FTP inputs to be editable when skipping Step4,
            # ensure _refresh_visibility or Step5 shows them (we earlier discussed that).
            self._goto_step(4)
            return

        # Default: Custom CLI Commands -> show Step4
        self._goto_step(3)

    def _step4_proceed(self):
        """
        Step4 → decide next navigation step.
        """

        # Read WLC commands
        try:
            if hasattr(self, "wlc_cmd_box"):
                self.wlc_cmds = [
                    l.strip() for l in self.wlc_cmd_box.toPlainText().splitlines() if l.strip()
                ]
        except Exception:
            self.wlc_cmds = getattr(self, "wlc_cmds", [])

        # Read AP commands
        try:
            if hasattr(self, "ap_cmd_box"):
                self.ap_cmds = [
                    l.strip() for l in self.ap_cmd_box.toPlainText().splitlines() if l.strip()
                ]
        except Exception:
            self.ap_cmds = getattr(self, "ap_cmds", [])

        # ---------------- WLC ONLY ----------------
        if self.operation_type == "WLC Only":

            if not self.wlc_cmds:
                QMessageBox.critical(self, "Missing", "Enter WLC Cmd List.")
                return

            self._fill_preview()
            self._goto_step(5)
            return

        # ---------------- AP ONLY ----------------
        if self.operation_type == "AP Only":

            # AP commands required only for Custom Cmd mode
            if getattr(self, "ap_mode", "") == "AP Custom Cmd List":
                if not self.ap_cmds:
                    QMessageBox.critical(self, "Missing", "Enter AP Cmd List.")
                    return

            if getattr(self, "ap_mode", "") == "AP Image Download":

                confirm = QMessageBox.warning(
                    self,
                    "Verify Image Before Proceeding",
                    "⚠️  Please double-check before continuing:\n\n"
                    "  • AP Model in your AP list file\n"
                    "  • Image filename matches that model\n\n"
                    "Wrong image will cause transfer failure.\n\n"
                    "Proceed?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )

                if confirm != QMessageBox.Yes:
                    return

            self._fill_preview()
            self._goto_step(5)
            return

        # ---------------- WLC & AP ----------------
        if self.operation_type == "WLC & AP":

            if not self.wlc_cmds:
                QMessageBox.critical(self, "Missing", "Enter WLC Cmd List.")
                return

            if not self.ap_cmds:
                QMessageBox.critical(self, "Missing", "Enter AP Cmd List.")
                return

            if getattr(self, "ap_mode", "") == "AP Image Download":

                confirm = QMessageBox.warning(
                    self,
                    "Verify Image Before Proceeding",
                    "⚠️  Please double-check before continuing:\n\n"
                    "  • AP Model from WLC AP summary\n"
                    "  • Image filename matches that model\n\n"
                    "Wrong image will cause transfer failure.\n\n"
                    "Proceed?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No
                )

                if confirm != QMessageBox.Yes:
                    return

            # go to Step5 Filters
            self._goto_step(4)

    def _step4_save(self):
        """
        Save WLC / AP command lists and FTP info (if AP Image Download).
        Defensive: checks widget existence before accessing them.
        """
        # Read WLC commands (if widget exists)
        wlc_cmds = []
        if hasattr(self, "wlc_cmd_box"):
            try:
                wlc_cmds = [l.strip() for l in self.wlc_cmd_box.toPlainText().splitlines() if l.strip()]
            except Exception:
                wlc_cmds = []

        # Read AP commands (if widget exists)
        ap_cmds = []
        if hasattr(self, "ap_cmd_box"):
            try:
                ap_cmds = [l.strip() for l in self.ap_cmd_box.toPlainText().splitlines() if l.strip()]
            except Exception:
                ap_cmds = []

        # Determine AP mode (prefer widget, fallback to state)
        ap_mode = None
        if hasattr(self, "ap_mode_dd") and callable(getattr(self.ap_mode_dd, "currentText", None)):
            ap_mode = self.ap_mode_dd.currentText()
        else:
            ap_mode = getattr(self, "ap_mode", "AP Custom Cmd List")

            # If AP Image Download is selected, validate FTP fields

            if not (ftp_user and ftp_user.text().strip()):
                ftp_missing = True
            if not (ftp_pasw and ftp_pasw.text()):
                ftp_missing = True

            if ftp_missing:
                QMessageBox.critical(self, "FTP Missing", "FTP(SFTP) fields are mandatory for AP Image Download.")
                return
            # If present, persist FTP into INI (defensive)
            if self.ini:
                try:
                    self.ini.bulk_set("FTP", {
                        "ftp_addr": ftp_addr.text().strip(),
                        "ftp_path": ftp_path.text().strip(),
                        "ftp_user(sftp)": ftp_user.text().strip(),
                        "ftp_pasw(sftp)": ftp_pasw.text(),
                        "scp_port": scp_port.text().strip() if scp_port and scp_port.text().strip() else "22"
                    })
                    self.ini.save()
                except Exception:
                    # non-fatal: continue but inform user
                    QMessageBox.warning(self, "Warning", "Failed to save FTP settings to config.ini (continuing).")

        # Persist ap_cmds/wlc_cmds to confd/ files
        try:
            os.makedirs(CONFD, exist_ok=True)
        except Exception:
            # if cannot create confd, warn and return
            QMessageBox.critical(self, "Disk Error", f"Unable to create config folder: {CONFD}")
            return

        # Save WLC commands (if any)
        if wlc_cmds:
            try:
                with open(os.path.join(CONFD, "cmdlist_wlc.txt"), "w", encoding="utf-8") as f:
                    f.write("\n".join(wlc_cmds) + "\n")
            except Exception as e:
                QMessageBox.warning(self, "Save Error", f"Failed to write WLC cmd list: {e}")

        # Save AP commands (if any) into file based on ap_device
        if ap_cmds:
            # decide filename; default to cos_qca
            fname = "cmdlist_cos_qca.txt"
            try:
                dev = getattr(self, "ap_device", "cos_qca")
                if dev == "cos":
                    fname = "cmdlist_cos.txt"
                elif dev == "cos_bcm":
                    fname = "cmdlist_cos_bcm.txt"
                else:
                    fname = "cmdlist_cos_qca.txt"
            except Exception:
                fname = "cmdlist_cos_qca.txt"

            try:
                with open(os.path.join(CONFD, fname), "w", encoding="utf-8") as f:
                    f.write("\n".join(ap_cmds) + "\n")
            except Exception as e:
                QMessageBox.warning(self, "Save Error", f"Failed to write AP cmd list: {e}")

        # Update internal state so the rest of the GUI knows the saved lists
        self.wlc_cmds = wlc_cmds
        self.ap_cmds = ap_cmds
        # Save SFTP credentials to INI so engine can read them
        if hasattr(self, "proto_dd") and self.ini:
            try:
                self.ini.bulk_set("FTP", {
                    "ftp_proto": self.proto_dd.currentText(),
                    "ftp_user": self.ftp_user.text().strip() if hasattr(self, "ftp_user") else "",
                    "ftp_pasw": self.ftp_pasw.text() if hasattr(self, "ftp_pasw") else "",
                })
                self.ini.save()
            except Exception:
                pass
        QMessageBox.information(self, "Saved", "Cmd lists (and FTP(SFTP) details if provided) saved under confd/")

    def _save_run_log(self):
        """
        Save the Run Log contents to data/WlanPoller_RunLog_YYYYMMDD_HHMMSS.txt.
        Defensive: works even if run_log widget is not present.
        """
        try:
            folder = DATA_DIR
            os.makedirs(folder, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            fn = os.path.join(folder, f"WlanPoller_RunLog_{ts}.txt")
            txt = ""
            if hasattr(self, "run_log") and self.run_log is not None:
                try:
                    txt = self.run_log.toPlainText()
                except Exception:
                    # fallback: read whatever attribute might exist
                    try:
                        txt = str(self.run_log)
                    except Exception:
                        txt = ""
            with open(fn, "w", encoding="utf-8") as f:
                f.write(txt)
            QMessageBox.information(self, "Saved", f"Run log saved to: {fn}")
        except Exception as e:
            # Show a friendly error if saving fails
            try:
                QMessageBox.warning(self, "Save Failed", f"Unable to save run log: {e}")
            except Exception:
                # last resort: print to console
                print("Unable to save run log:", e)

    def _on_ap_mode_changed(self, text: str):
        self.ap_mode = text
        is_image = (text == "AP Image Download")

        # cmd box always visible when AP is involved
        if hasattr(self, "ap_cmd_box"):
            self.ap_cmd_box.setVisible(True)

        # ftp_group only visible when Image Download selected
        if hasattr(self, "ftp_group"):
            self.ftp_group.setVisible(is_image)

        self._refresh_visibility()
    def _export_ap_table(self):
        """
        Export the AP Table to an Excel file under data/.
        Defensive: checks for openpyxl and widget presence.
        """
        if Workbook is None:
            QMessageBox.warning(self, "Missing", "openpyxl is not installed. Run: pip install openpyxl")
            return

        # Quick helper: safely read cell text
        def _cell_text(table, r, c):
            try:
                it = table.item(r, c)
                return it.text() if (it and it.text()) else ""
            except Exception:
                return ""

        # Build rows from AP table (single pass)
        rows = []
        if hasattr(self, "ap_table"):
            try:
                for r in range(self.ap_table.rowCount()):
                    # Expect columns: 0=AP Name, 1=AP IP, 2=AP Model, 3=Status
                    # If the table doesn't have 4 columns yet, we still read safely
                    # when building rows for export (single pass)
                    name = _cell_text(self.ap_table, r, 0)
                    model = _cell_text(self.ap_table, r, 1)
                    ip = _cell_text(self.ap_table, r, 2)
                    status = _cell_text(self.ap_table, r, 3)
                    rows.append([name, model, ip, status])
                    # and set headers accordingly: ["AP Name","AP Model","AP IP","Status"]

            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed reading AP table: {e}")
                return
        else:
            QMessageBox.information(self, "No Data", "AP Table not found / empty.")
            return

        # Prepare Excel file
        folder = DATA_DIR
        os.makedirs(folder, exist_ok=True)
        fn = os.path.join(folder, f"WlanPoller_AP_Table_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "AP Results"

            headers = ["AP Name", "AP IP", "AP Model", "Status"]
            ws.append(headers)

            # header styling
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = XLFont(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # write rows (one append per row is fine)
            for rdata in rows:
                ws.append(rdata)

            # freeze & autofilter
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

            # nice column widths
            ws.column_dimensions["A"].width = 32  # AP Name
            ws.column_dimensions["B"].width = 18  # AP IP
            ws.column_dimensions["C"].width = 18  # AP Model
            ws.column_dimensions["D"].width = 80  # Status

            wb.save(fn)
            QMessageBox.information(self, "Excel Exported", f"Saved: {fn}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Failed to save Excel file: {e}")

    def _export_vuln_table(self):
        """
        Export the Vulnerable APs table to an Excel file under data/.
        Defensive: checks for openpyxl and widget presence.
        """
        if Workbook is None:
            QMessageBox.warning(self, "Missing", "openpyxl is not installed. Run: pip install openpyxl")
            return

        # Build rows from Vulnerable table
        rows = []
        if hasattr(self, "vuln_table"):
            try:
                for r in range(self.vuln_table.rowCount()):
                    name = self.vuln_table.item(r, 0).text() if self.vuln_table.item(r, 0) else ""
                    model = self.vuln_table.item(r, 1).text() if self.vuln_table.item(r, 1) else ""
                    ip = self.vuln_table.item(r, 2).text() if self.vuln_table.item(r, 2) else ""
                    recovery = self.vuln_table.item(r, 3).text() if self.vuln_table.item(r, 3) else ""
                    rows.append([name, model, ip, recovery])
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed reading Vulnerable table: {e}")
                return
        else:
            QMessageBox.information(self, "No Data", "Vulnerable APs table not found / empty.")
            return

        # Prepare Excel file
        folder = DATA_DIR
        os.makedirs(folder, exist_ok=True)
        fn = os.path.join(folder, f"WlanPoller_Susceptible_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Susceptible APs"

            headers = ["AP Name", "AP Model", "AP IP", "Recovery"]
            ws.append(headers)

            # header styling
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=c)
                cell.font = XLFont(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # write rows
            for rdata in rows:
                ws.append(rdata)

            # freeze & autofilter
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

            # nice column widths
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 90

            wb.save(fn)
            QMessageBox.information(self, "Excel Exported", f"Saved: {fn}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Failed to save Excel file: {e}")

    def _open_data_folder(self):

        folder = str(DATA_DIR)


        # Prefer last run folder if available
        if hasattr(self, "last_status_file"):
            try:
                folder = os.path.dirname(self.last_status_file)
            except Exception:
                pass

        try:
            os.makedirs(folder, exist_ok=True)

            if sys.platform.startswith("win"):
                os.startfile(folder)
            elif sys.platform == "darwin":
                os.system(f'open "{folder}"')
            else:
                os.system(f'xdg-open "{folder}"')

        except Exception as e:
            QMessageBox.warning(self, "Open Folder Failed", str(e))
    def _on_workflow_change(self, v: str):
        self.workflow = v

    # ... (remaining methods: _step3_proceed, _on_ap_mode_changed, _step4_save, _step4_proceed, _enforce_one_filter,
    # _update_ap_device_from_model, _step5_preview are defined earlier in file - kept unchanged for brevity) ...
    # For completeness they are implemented above in the full content.
    def _on_proto_changed(self, proto):
        is_sftp = (proto == "SFTP")

        # Server IP and Remote Path — only needed for SFTP
        for attr in ("ftp_addr", "ftp_path", "ftp_user", "ftp_pasw",
                     "ftp_user_label", "ftp_pasw_label"):
            if hasattr(self, attr):
                getattr(self, attr).setVisible(is_sftp)

        # Also hide/show the form labels via the ftp_layout rows
        if hasattr(self, "ftp_group") and self.ftp_group.layout():
            lay = self.ftp_group.layout()
            for i in range(lay.rowCount()):
                label_item = lay.itemAt(i, QFormLayout.LabelRole)
                if label_item and label_item.widget():
                    txt = label_item.widget().text()
                    if txt in ("Server IP:", "Remote Path:"):
                        label_item.widget().setVisible(is_sftp)
    def _enforce_one_filter(self, _):
        if self.chk_site.isChecked() and self.chk_model.isChecked():
            sender = self.sender()
            if sender == self.chk_site:
                self.chk_model.setChecked(False)
            else:
                self.chk_site.setChecked(False)

    def _update_ap_device_from_model(self, label: str):
        if label in ("All AP Models", "AP1852/2802/3802/4802"):
            self.ap_device = "cos"
        elif label == "C9105AX/9115AX/9120AX":
            self.ap_device = "cos_bcm"
        else:
            self.ap_device = "cos_qca"

    def _step5_preview(self):

        # ---- WLC Only → filters not allowed ----
        if self.operation_type == "WLC Only":
            self.ap_filter_mode = "NONE"
            self._fill_preview()
            self._goto_step(5)
            return

        # ---- Normal filtering ----
        if self.chk_site.isChecked():
            self.ap_filter_mode = "SITE"
            self.site_tag = self.site_tag_txt.text().strip()
            if not self.site_tag:
                QMessageBox.critical(self, "Missing", "Enter SiteTag Name.")
                return

        elif self.chk_model.isChecked():
            self.ap_filter_mode = "MODEL"
            self.model_group = self.model_dd.currentText()

        else:
            self.ap_filter_mode = "NONE"

        self._fill_preview()
        self._goto_step(5)

    def _on_finished(self, summary: dict):
        try:
            # ---------------- BASIC INFO ----------------
            start = summary.get("start")
            end = summary.get("end")
            elapsed = int((end - start).total_seconds()) if start and end else 0

            lines = []
            lines.append(f"Operation Type Selected in Step1: {summary.get('operation', '')}")

            if summary.get("operation") != "AP Only":
                wlc_ip = self.wlc_ip.text().strip() if hasattr(self, "wlc_ip") else ""
                if wlc_ip:
                    lines.append(f"WLC IP address: {wlc_ip}")

            if summary.get("operation") in ("WLC & AP", "AP Only"):
                lines.append(f"Total number of APs Processed: {summary.get('ap_total', 0)}")
                lines.append(f"Success APs: {summary.get('ap_success', 0)}")
                lines.append(f"Failed APs: {summary.get('ap_failed', 0)}")

            if summary.get("SiteTagNameFilter"):
                lines.append(
                    f"Selected {summary.get('ApFilteredCnt', 0)} out of {summary.get('TotalApCnt', 0)} total Aps from Site Tag filter '{summary.get('SiteTagNameFilter')}'"
                )

            lines.append(f"Time taken: {elapsed}s")

            if summary.get("operation") == "WLC Only":
                out = summary.get("wlc_output", "")
                lines.append(
                    f"writing outputs to the folder 'data' to the file named {os.path.basename(out) if out else 'eWLC-9800_outputs.txt'}"
                )
            else:
                lines.append(f"Outputs stored in: {summary.get('data_dir', '')}")

            status_file = summary.get("status_summary_file", "")
            if status_file:
                lines.append(f"Status Check Summary File: {status_file}")

            if hasattr(self, "results_summary"):
                try:
                    self.results_summary.setPlainText("\n".join(lines))
                except Exception:
                    pass

            # ---------------- FLAGS ----------------
            wf = summary.get("workflow") or getattr(self, "workflow", None)
            is_wlc_only = summary.get("operation") == "WLC Only"

            # ---------------- VULNERABLE TABLE ----------------
            if hasattr(self, "vuln_table"):
                try:
                    self.vuln_table.setRowCount(0)
                except Exception:
                    pass

            if (not is_wlc_only) and wf == "AP Flash Checker":
                if hasattr(self, "run_log"):
                    try:
                        self.run_log.append("\n===== PARSER / FLASH CHECK SUMMARY =====")
                    except Exception:
                        pass

                vuln_rows = summary.get("vulnerable_rows", [])
                # Update AP table with correct model from parser

                vuln_count = len(vuln_rows)

                if hasattr(self, "run_log"):
                    self.run_log.append(f"Total Susceptible APs Detected: {vuln_count}")
                if hasattr(self, "vuln_table"):
                    try:
                        for vr in vuln_rows:
                            r = self.vuln_table.rowCount()
                            self.vuln_table.insertRow(r)
                            self.vuln_table.setItem(r, 0, QTableWidgetItem(vr.get("ap_name", "")))
                            self.vuln_table.setItem(r, 1, QTableWidgetItem(vr.get("ap_model", "")))
                            self.vuln_table.setItem(r, 2, QTableWidgetItem(vr.get("ap_ip", "")))
                            self.vuln_table.setItem(r, 3, QTableWidgetItem(vr.get("recovery", "")))
                    except Exception as e:
                        if hasattr(self, "run_log"):
                            try:
                                self.run_log.append(f"[DEBUG] Failed populating vuln_table: {e}")
                            except Exception:
                                pass

            # ---------------- AP TABLE (skip for WLC only) ----------------
            if not is_wlc_only:
                try:
                    # So AP Only mode never loads old WLC files.
                    need_populate = (
                            hasattr(self, "ap_table")
                            and self.ap_table.rowCount() == 0

                    )

                    if need_populate:
                        cand_files = []
                        try:
                            for fn in os.listdir(CONFD):
                                if fn.lower().startswith("ap_ip_list") and fn.lower().endswith(".txt"):
                                    cand_files.append(os.path.join(CONFD, fn))
                        except Exception:
                            cand_files = []

                        src_file = ""
                        for f in cand_files:
                            if "all" in os.path.basename(f).lower():
                                src_file = f
                                break
                        if not src_file and cand_files:
                            src_file = cand_files[0]

                        if src_file and os.path.exists(src_file):
                            try:
                                if src_file and os.path.exists(src_file):
                                    try:
                                        with open(src_file, "r", encoding="utf-8", errors="ignore") as fh:
                                            # Clear table and populate rows in new order:
                                            # AP Name | AP Model | AP IP | Status
                                            if hasattr(self, "ap_table"):
                                                self.ap_table.setRowCount(0)
                                            rows_added = 0
                                            for i, line in enumerate(fh):
                                                if i >= 200:
                                                    break
                                                s = line.strip()
                                                if not s:
                                                    continue

                                                # split by comma if present (CSV), otherwise by whitespace
                                                parts = [p.strip() for p in (s.split(",") if "," in s else s.split())]

                                                # Basic defaults
                                                ip = parts[0] if len(parts) >= 1 else ""
                                                model = "-"
                                                name = ""

                                                # If line has 3+ tokens assume: ip model name
                                                if len(parts) >= 3:
                                                    ip = parts[0]
                                                    model = parts[1]
                                                    # name may contain spaces when comma-separated; join remainder for safety
                                                    name = " ".join(parts[2:]).strip()
                                                elif len(parts) == 2:
                                                    # ambiguous two-column line — prefer CSV interpretation (ip,name),
                                                    # otherwise treat second token as name and leave model unknown.
                                                    name = parts[1]

                                                # Save mapping so later ap_update can use it
                                                try:
                                                    if name:
                                                        self.ap_name_map[ip] = name
                                                except Exception:
                                                    pass

                                                # Insert into table as: Name | Model | IP | Status
                                                if hasattr(self, "ap_table"):
                                                    r = self.ap_table.rowCount()
                                                    self.ap_table.insertRow(r)
                                                    try:
                                                        self.ap_table.setItem(r, 0, QTableWidgetItem(name))
                                                        self.ap_table.setItem(r, 1, QTableWidgetItem(model))
                                                        self.ap_table.setItem(r, 2, QTableWidgetItem(ip))
                                                        self.ap_table.setItem(r, 3, QTableWidgetItem("Pending"))
                                                    except Exception:
                                                        # fallback: attempt safer sets
                                                        try:
                                                            if not self.ap_table.item(r, 0):
                                                                self.ap_table.setItem(r, 0, QTableWidgetItem(name))
                                                        except Exception:
                                                            pass

                                                rows_added += 1

                                            if hasattr(self, "run_log"):
                                                try:
                                                    self.run_log.append(
                                                        f"[DEBUG] Populated AP Table with {rows_added} rows from {os.path.basename(src_file)}")
                                                except Exception:
                                                    pass
                                    except Exception as e:
                                        if hasattr(self, "run_log"):
                                            try:
                                                self.run_log.append(f"[DEBUG] Exception reading fallback AP list: {e}")
                                            except Exception:
                                                pass

                                            except Exception:
                                                # fallback: attempt safer sets
                                                try:
                                                    if not self.ap_table.item(r, 0): self.ap_table.setItem(r, 0,
                                                                                                           QTableWidgetItem(
                                                                                                               name))
                                                except Exception:
                                                    pass
                                        rows_added += 1

                                if hasattr(self, "run_log"):
                                    try:
                                        self.run_log.append(
                                            f"[DEBUG] Populated AP Table with {rows_added} rows from {os.path.basename(src_file)}")
                                    except Exception:
                                        pass
                            except Exception as e:
                                if hasattr(self, "run_log"):
                                    try:
                                        self.run_log.append(f"[DEBUG] Exception reading fallback AP list: {e}")
                                    except Exception:
                                        pass

                except Exception as e:
                    if hasattr(self, "run_log"):
                        try:
                            self.run_log.append(f"[DEBUG] AP table population error: {e}")
                        except Exception:
                            pass
            # ---------------- FINAL MODEL CORRECTION ----------------
            vuln_rows = summary.get("vulnerable_rows", [])

            for vr in vuln_rows:
                ip = vr.get("ap_ip")
                model = vr.get("ap_model")

                for r in range(self.ap_table.rowCount()):
                    ip_item = self.ap_table.item(r, 2)
                    if ip_item and ip_item.text() == ip:
                        if model and model != "UNKNOWN":
                            self.ap_table.setItem(r, 1, QTableWidgetItem(model))
                        break


            # ---------------- UNLOCK UI ----------------
            if hasattr(self, "sidebar"):
                try:
                    self.sidebar.setEnabled(True)
                except Exception:
                    pass

            # Re-enable Close button if present
            if hasattr(self, "btn_close"):
                try:
                    self.btn_close.setEnabled(True)
                except Exception:
                    pass
            if status_file:
                self.last_status_file = summary.get("status_summary_file", "")

        except Exception as e:
            if hasattr(self, "run_log"):
                try:
                    self.run_log.append(f"[ERROR] _on_finished exception: {e}")
                except Exception:
                    pass
            else:
                print("[ERROR] _on_finished exception:", e)

        # finally mark run_in_progress false
        try:
            self.run_count += 1
            if self.run_count >= 2:
                QMessageBox.information(
                    self,
                    "Restart Recommended",
                    "You have completed multiple runs in the same session.\n\n"
                    "For best results, please save your logs and restart the application before running again.\n\n"
                    "Continuing without restart may cause unexpected behaviour."
                )
            self.run_in_progress = False

        except Exception:
            pass

    def _on_ap_update(self, *args):
        try:
            if len(args) < 4:
                return

            ip = str(args[1]) if args[1] else ""
            model = str(args[2]) if args[2] else ""
            status = str(args[3]) if args[3] else ""
            name = str(args[4]) if len(args) > 4 and args[4] else ""

            if not hasattr(self, "ap_table"):
                return

            # Find row by IP instead of trusting index
            target_row = -1
            for r in range(self.ap_table.rowCount()):
                item = self.ap_table.item(r, 2)  # AP IP column
                if item and item.text() == ip:
                    target_row = r
                    break

            # If IP not found, append new row
            if target_row == -1:
                target_row = self.ap_table.rowCount()
                self.ap_table.insertRow(target_row)

            # Fallback name logic
            if not name and hasattr(self, "ap_name_map"):
                name = self.ap_name_map.get(ip, "")

            self.ap_table.setItem(target_row, 0, QTableWidgetItem(name))
            # Prevent overwriting correct model with UNKNOWN
            existing_item = self.ap_table.item(target_row, 1)
            existing_model = existing_item.text() if existing_item else ""

            if model and model != "UNKNOWN":
                final_model = model
            elif existing_model and existing_model != "UNKNOWN":
                final_model = existing_model
            else:
                final_model = "UNKNOWN"

            self.ap_table.setItem(target_row, 1, QTableWidgetItem(final_model))
            self.ap_table.setItem(target_row, 2, QTableWidgetItem(ip))
            self.ap_table.setItem(target_row, 3, QTableWidgetItem(status))
            self.ap_table.resizeRowToContents(target_row)
            if hasattr(self, "run_log"):
                self.run_log.append(f"[AP_UPDATE] ip={ip} status={status}")

        except Exception as e:
            if hasattr(self, "run_log"):
                self.run_log.append(f"[ERROR] _on_ap_update exception: {e}")

    def _fill_preview(self):
        """
        Build Step-6 Preview safely.
        Never throws exceptions.
        """

        def get_txt(widget):
            try:
                return widget.text().strip()
            except Exception:
                return ""

        def get_dd(widget):
            try:
                return widget.currentText().strip()
            except Exception:
                return ""

        lines = []
        sec = 1

        def section(title):
            nonlocal sec
            lines.append(f"{sec}) {title}")
            sec += 1

        # ---------------- Operation ----------------
        section("Operation Type")

        op = getattr(self, "operation_type", "")
        lines.append(f"   - {op}")

        if op != "AP Only" and hasattr(self, "wlc_ip"):
            lines.append(f"   - WLC IP: {get_txt(self.wlc_ip)}")

        # ---------------- Credentials ----------------
        lines.append("")
        section("Credentials")

        if op != "AP Only" and hasattr(self, "wlc_user"):
            lines.append(f"   - WLC Username: {get_txt(self.wlc_user)}")

        if op != "WLC Only" and hasattr(self, "ap_user"):
            lines.append(f"   - AP Username: {get_txt(self.ap_user)}")

        # ---------------- Workflow ----------------
        # ---------------- Workflow ----------------
        lines.append("")
        section("Workflow")

        op = getattr(self, "operation_type", "")

        # For WLC Only operation → fixed workflow
        if op == "WLC Only":
            wf = "Custom WLC CLI Commands"
        else:
            wf = getattr(self, "workflow", "")
            if hasattr(self, "workflow_dd"):
                wf = get_dd(self.workflow_dd)

        lines.append(f"   - {wf}")

        # ---------------- CLI COMMANDS  (MERGED) ----------------
        wlc_cmds = []
        ap_cmds = []

        if op != "AP Only":
            wlc_cmds = getattr(self, "wlc_cmds", [])
            if not wlc_cmds and hasattr(self, "wlc_cmd_box"):
                wlc_cmds = [c.strip() for c in self.wlc_cmd_box.toPlainText().splitlines() if c.strip()]

        if op != "WLC Only":
            ap_cmds = getattr(self, "ap_cmds", [])
            if not ap_cmds and hasattr(self, "ap_cmd_box"):
                ap_cmds = [c.strip() for c in self.ap_cmd_box.toPlainText().splitlines() if c.strip()]

        if wlc_cmds or ap_cmds:
            lines.append("")
            section("CLI Commands")

            if wlc_cmds:
                lines.append("   WLC:")
                for c in wlc_cmds:
                    lines.append(f"      • {c}")

            if ap_cmds:
                lines.append("   AP:")
                for c in ap_cmds:
                    lines.append(f"      • {c}")


        # ---------------- AP Filters ----------------
        lines.append("")
        section("AP Filters")

        if op == "WLC Only":
            lines.append("   - Not Applicable")
        else:
            mode = getattr(self, "ap_filter_mode", "NONE")

            if mode == "SITE":
                site = getattr(self, "site_tag", "")
                if hasattr(self, "site_tag_txt"):
                    site = get_txt(self.site_tag_txt) or site
                lines.append(f"   - Site Tag: {site}")

            elif mode == "MODEL":
                model = getattr(self, "model_group", "")
                if hasattr(self, "model_dd"):
                    model = get_dd(self.model_dd)
                lines.append(f"   - Model Group: {model}")

            else:
                lines.append("   - No Filters Applied")

        # ---------------- Render Preview ----------------
        if hasattr(self, "preview_text"):
            try:
                self.preview_text.setPlainText("\n".join(lines))
            except Exception:
                pass

    def _run_parser(self):
        """
        Run parser/search over the most recent 'data' leaf folder.
        Supports regex (preferred) or plain substring if regex compilation fails.
        Only searches files appropriate for the selected parser mode:
          - "WLC files" => only files ending with "_outputs.txt"
          - "AP files"  => everything except files ending with "_outputs.txt"
        Results are written to self.parser_out (QTextEdit).
        Defensive: checks attributes before using them.
        """
        try:
            # get pattern text
            if not hasattr(self, "parser_pattern") or self.parser_pattern is None:
                QMessageBox.critical(self, "Missing", "Parser input control not found.")
                return
            pat_text = self.parser_pattern.text().strip()
            if not pat_text:
                QMessageBox.critical(self, "Missing", "Enter a pattern to search.")
                return

            # determine latest data folder
            data_root = DATA_DIR
            if not os.path.exists(data_root):
                # nothing to search
                out = "No 'data' folder found."
                if hasattr(self, "parser_out"):
                    self.parser_out.setPlainText(out)
                else:
                    print(out)
                return

            leafs = []
            for top, dirs, files in os.walk(data_root):
                if files:
                    leafs.append(top)
            latest = max(leafs) if leafs else data_root

            # choose mode
            use_wlc = False
            if hasattr(self, "parser_mode") and callable(getattr(self.parser_mode, "currentText", None)):
                use_wlc = (self.parser_mode.currentText() == "WLC files")
            else:
                # fallback if parser_mode is stored somewhere else
                use_wlc = getattr(self, "parser_mode_value", "WLC files") == "WLC files"

            out_lines = [f"Searching in: {latest}", f"Mode: {'WLC files' if use_wlc else 'AP files'}",
                         f"Pattern: {pat_text}", ""]

            # try compile regex; if it fails, fall back to substring search but report the regex error
            rx = None
            try:
                rx = re.compile(pat_text)
            except Exception as e:
                out_lines.append(
                    f"[Note] Pattern is not a valid regex — falling back to substring search. Regex error: {e}")
                out_lines.append("")

            found_any = False
            try:
                files = sorted(os.listdir(latest))
            except Exception as e:
                out_lines.append(f"[ERROR] Unable to list files in {latest}: {e}")
                if hasattr(self, "parser_out"):
                    self.parser_out.setPlainText("\n".join(out_lines))
                return

            for fn in files:
                # filter files depending on mode
                if use_wlc and not fn.endswith("_outputs.txt"):
                    continue
                if (not use_wlc) and fn.endswith("_outputs.txt"):
                    continue

                fp = os.path.join(latest, fn)
                try:
                    with open(fp, "r", encoding="utf-8", errors="ignore") as f:
                        txt = f.read()
                    match = False
                    if rx:
                        if rx.search(txt):
                            match = True
                    else:
                        if pat_text in txt:
                            match = True
                    if match:
                        found_any = True
                        out_lines.append(f"[MATCH] {fn}")
                except Exception as e:
                    out_lines.append(f"[ERROR] {fn}: {e}")

            if not found_any:
                out_lines.append("No matches found.")

            # write to parser_out if present
            if hasattr(self, "parser_out") and self.parser_out is not None:
                self.parser_out.setPlainText("\n".join(out_lines))
            else:
                print("\n".join(out_lines))

        except Exception as e:
            # last-resort error handling
            try:
                QMessageBox.critical(self, "Parser Error", f"Unexpected error while parsing: {e}")
            except Exception:
                print("Unexpected error in _run_parser:", e)

    def _refresh_visibility(self):

        # ---------------- STEP 1 ----------------
        if hasattr(self, "ap_upload_row"):
            self.ap_upload_row.setVisible(self.operation_type == "AP Only")
        elif hasattr(self, "ap_upload_box"):
            self.ap_upload_box.setVisible(self.operation_type == "AP Only")

        if hasattr(self, "btn_step1_next"):
            if self.operation_type == "AP Only":
                self.btn_step1_next.setEnabled(
                    bool(getattr(self, "ap_list_file", "")) or
                    bool(getattr(self, "ap_list_path", ""))
                )
            else:
                self.btn_step1_next.setEnabled(True)

        # ---------------- STEP 2 (FIXED) ----------------
        show_wlc = self.operation_type in ("WLC Only", "WLC & AP")
        show_ap = self.operation_type in ("AP Only", "WLC & AP")

        if hasattr(self, "wlc_block"):
            self.wlc_block.setVisible(show_wlc)

        if hasattr(self, "ap_block"):
            self.ap_block.setVisible(show_ap)

        # ---------------- STEP 4 (FIXED) ----------------
        if hasattr(self, "ap_cmd_section"):
            self.ap_cmd_section.setVisible(self.operation_type != "WLC Only")

        # ---------------- COMMAND BOX VISIBILITY ----------------
        ap_mode_text = (
            self.ap_mode_dd.currentText()
            if hasattr(self, "ap_mode_dd")
            else getattr(self, "ap_mode", "AP Custom Cmd List")
        )

        show_wlc_cmd = (
                self.operation_type in ("WLC Only", "WLC & AP")
                and getattr(self, "workflow", "") != "AP Flash Checker"
        )
        if hasattr(self, "wlc_cmd_section"):
            self.wlc_cmd_section.setVisible(show_wlc_cmd)
        elif hasattr(self, "wlc_cmd_box"):
            self.wlc_cmd_box.setVisible(show_wlc_cmd)
        show_ap_cmd = (
                              self.operation_type in ("AP Only", "WLC & AP")
                      ) or getattr(self, "workflow", "") == "AP Flash Checker"
        if hasattr(self, "ap_cmd_box"):
            self.ap_cmd_box.setVisible(show_ap_cmd)
        # ---------------- FTP ----------------
        ftp_visible = (
                self.operation_type in ("AP Only", "WLC & AP")
                and ap_mode_text == "AP Image Download"
        )
        if hasattr(self, "ftp_group"):
            self.ftp_group.setVisible(ftp_visible)

        # ---------------- FILTERS ----------------
        filters_allowed = self.operation_type in ("WLC & AP", "AP Only")
        # Disable model filter for AP Flash Checker in WLC & AP mode
        if (
                self.operation_type == "WLC & AP"
                and getattr(self, "workflow", "") == "AP Flash Checker"
        ):
            if hasattr(self, "chk_model"):
                self.chk_model.setChecked(False)
                self.chk_model.setVisible(False)

            if hasattr(self, "model_dd"):
                self.model_dd.setVisible(False)
        else:
            # Restore visibility for other workflows
            if hasattr(self, "chk_model"):
                self.chk_model.setVisible(True)

            if hasattr(self, "model_dd"):
                self.model_dd.setVisible(True)
        if hasattr(self, "chk_site") and hasattr(self, "chk_model"):
            if not filters_allowed:
                self.chk_site.setChecked(False)
                self.chk_model.setChecked(False)

        # Force Qt layout recalculation
        if self.centralWidget():
            self.centralWidget().updateGeometry()

    def _post_init_layout_fix(self):
        # first apply visibility rules
        self._refresh_visibility()

        # go to step0 to ensure consistent base layout
        self.stack.setCurrentIndex(0)

        # allow Qt to finish geometry

    def _ui_progress_update(self, pct):
        if hasattr(self, "progress"):
            self.progress.setValue(pct)

    def changeEvent(self, event):
        from PySide6.QtCore import QEvent

        if event.type() == QEvent.WindowStateChange:
            if not self.isMinimized():
                if hasattr(self, "ap_table"):
                    self.ap_table.resizeRowsToContents()

        super().changeEvent(event)


def resource_path(relpath: str) -> str:
    """
    Return a filesystem path to `relpath` that works when running normally
    and when bundled by PyInstaller.
    Example: resource_path("assets/cisco_logo.ico")
    """
    if getattr(sys, "frozen", False):
        # Running in a PyInstaller bundle
        base = getattr(sys, "_MEIPASS", os.path.abspath(os.path.dirname(__file__)))
    else:
        base = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base, relpath)


def main():
    app = QApplication(sys.argv)
    apply_global_style(app)
    app.setWindowIcon(QIcon(resource_path("assets/ciscologo.ico")))
    win = MainWindow()
    # ensure worker stop when the app quits (in case user closes app from other places)
    app.aboutToQuit.connect(win._stop_worker)
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
